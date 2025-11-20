"""
Utilidades para exportar reportes a PDF y Excel
"""
from datetime import datetime, date
from typing import List, Dict, Any
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class ExportadorReportes:
    """Clase para exportar reportes a diferentes formatos"""
    
    @staticmethod
    def exportar_a_pdf(datos: Dict[str, Any], tipo_reporte: str, nombre_archivo: str = None) -> str:
        """
        Exporta un reporte a PDF
        
        Args:
            datos: Diccionario con los datos del reporte
            tipo_reporte: Tipo de reporte (cierre_diario, cierre_mensual, etc.)
            nombre_archivo: Nombre del archivo (opcional)
            
        Returns:
            Ruta del archivo generado
        """
        if not nombre_archivo:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"reporte_{tipo_reporte}_{timestamp}.pdf"
        
        # Crear directorio de reportes si no existe
        os.makedirs("reportes", exist_ok=True)
        ruta_completa = os.path.join("reportes", nombre_archivo)
        
        # Crear documento PDF
        doc = SimpleDocTemplate(ruta_completa, pagesize=letter)
        elementos = []
        estilos = getSampleStyleSheet()
        
        # Estilo para el título
        estilo_titulo = ParagraphStyle(
            'CustomTitle',
            parent=estilos['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#262262'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Estilo para subtítulos
        estilo_subtitulo = ParagraphStyle(
            'CustomSubtitle',
            parent=estilos['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#666666'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        # Generar contenido según el tipo de reporte
        if tipo_reporte == "cierre_diario":
            ExportadorReportes._generar_pdf_cierre_diario(elementos, datos, estilo_titulo, estilo_subtitulo)
        elif tipo_reporte == "cierre_mensual":
            ExportadorReportes._generar_pdf_cierre_mensual(elementos, datos, estilo_titulo, estilo_subtitulo)
        elif tipo_reporte == "compras_periodo":
            ExportadorReportes._generar_pdf_compras(elementos, datos, estilo_titulo, estilo_subtitulo)
        elif tipo_reporte == "productos_existencias":
            ExportadorReportes._generar_pdf_productos(elementos, datos, estilo_titulo, estilo_subtitulo)
        elif tipo_reporte == "cartera_clientes":
            ExportadorReportes._generar_pdf_clientes(elementos, datos, estilo_titulo, estilo_subtitulo)
        elif tipo_reporte == "cartera_proveedores":
            ExportadorReportes._generar_pdf_proveedores(elementos, datos, estilo_titulo, estilo_subtitulo)
        elif tipo_reporte == "cartera_empleados":
            ExportadorReportes._generar_pdf_empleados(elementos, datos, estilo_titulo, estilo_subtitulo)
        
        # Generar PDF
        doc.build(elementos)
        return ruta_completa
    
    @staticmethod
    def _generar_pdf_cierre_diario(elementos, datos, estilo_titulo, estilo_subtitulo):
        """Genera PDF para cierre de caja diario"""
        resumen = datos.get('resumen', {})
        ventas = datos.get('ventas', [])
        fecha = datos.get('fecha')
        
        # Título
        elementos.append(Paragraph(f"Cierre de Caja Diario", estilo_titulo))
        elementos.append(Paragraph(f"Fecha: {fecha.strftime('%d/%m/%Y') if hasattr(fecha, 'strftime') else fecha}", estilo_subtitulo))
        elementos.append(Spacer(1, 0.3*inch))
        
        # Resumen
        datos_resumen = [
            ['Métrica', 'Valor'],
            ['Total Ventas', str(resumen.get('total_ventas', 0))],
            ['Total Ingresos', f"Q{resumen.get('total_ingresos', 0):,.2f}"],
            ['Efectivo', f"Q{resumen.get('efectivo', 0):,.2f}"],
            ['Tarjeta', f"Q{resumen.get('tarjeta', 0):,.2f}"],
            ['Transferencia', f"Q{resumen.get('transferencia', 0):,.2f}"],
        ]
        
        tabla_resumen = Table(datos_resumen, colWidths=[3*inch, 2*inch])
        tabla_resumen.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#262262')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elementos.append(tabla_resumen)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Detalle de ventas
        if ventas:
            elementos.append(Paragraph("Detalle de Ventas", estilo_subtitulo))
            elementos.append(Spacer(1, 0.2*inch))
            
            datos_ventas = [['Factura', 'Cliente', 'Método Pago', 'Total']]
            for v in ventas:
                datos_ventas.append([
                    v.get('numero_factura', ''),
                    v.get('cliente', ''),
                    v.get('metodo_pago', '').capitalize(),
                    f"Q{v.get('total', 0):,.2f}"
                ])
            
            tabla_ventas = Table(datos_ventas, colWidths=[1.5*inch, 2*inch, 1.5*inch, 1.5*inch])
            tabla_ventas.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#262262')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elementos.append(tabla_ventas)
    
    @staticmethod
    def _generar_pdf_cierre_mensual(elementos, datos, estilo_titulo, estilo_subtitulo):
        """Genera PDF para cierre de caja mensual"""
        resumen = datos.get('resumen', {})
        datos_diarios = datos.get('datos_diarios', [])
        
        meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        
        elementos.append(Paragraph(f"Cierre de Caja Mensual", estilo_titulo))
        elementos.append(Paragraph(f"{meses[datos['mes']]} {datos['año']}", estilo_subtitulo))
        elementos.append(Spacer(1, 0.3*inch))
        
        # Resumen
        datos_resumen = [
            ['Métrica', 'Valor'],
            ['Total Ventas', str(resumen.get('total_ventas', 0))],
            ['Total Ingresos', f"Q{resumen.get('total_ingresos', 0):,.2f}"],
            ['Promedio por Venta', f"Q{resumen.get('promedio_venta', 0):,.2f}"],
        ]
        
        tabla_resumen = Table(datos_resumen, colWidths=[3*inch, 2*inch])
        tabla_resumen.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#262262')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elementos.append(tabla_resumen)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Detalle por día
        if datos_diarios:
            datos_tabla = [['Fecha', 'Ventas', 'Efectivo', 'Tarjeta', 'Total']]
            for d in datos_diarios:
                fecha = d.get('fecha')
                if isinstance(fecha, str):
                    try:
                        fecha = datetime.strptime(fecha, "%Y-%m-%d").date()
                    except:
                        pass
                
                datos_tabla.append([
                    fecha.strftime("%d/%m/%Y") if hasattr(fecha, 'strftime') else str(fecha),
                    str(d.get('total_ventas', 0)),
                    f"Q{d.get('efectivo', 0):,.2f}",
                    f"Q{d.get('tarjeta', 0):,.2f}",
                    f"Q{d.get('total_ingresos', 0):,.2f}"
                ])
            
            tabla = Table(datos_tabla, colWidths=[1.3*inch, 1*inch, 1.3*inch, 1.3*inch, 1.3*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#262262')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elementos.append(tabla)
    
    @staticmethod
    def _generar_pdf_compras(elementos, datos, estilo_titulo, estilo_subtitulo):
        """Genera PDF para compras por periodo"""
        resumen = datos.get('resumen', {})
        compras = datos.get('compras', [])
        
        elementos.append(Paragraph(f"Compras por Periodo", estilo_titulo))
        elementos.append(Paragraph(
            f"{datos['fecha_inicio'].strftime('%d/%m/%Y')} - {datos['fecha_fin'].strftime('%d/%m/%Y')}",
            estilo_subtitulo
        ))
        elementos.append(Spacer(1, 0.3*inch))
        
        # Resumen
        datos_resumen = [
            ['Total Compras', str(resumen.get('total_compras', 0))],
            ['Total Gastado', f"Q{resumen.get('total_gastado', 0):,.2f}"],
            ['Promedio por Compra', f"Q{resumen.get('promedio_compra', 0):,.2f}"],
        ]
        
        tabla_resumen = Table(datos_resumen, colWidths=[3*inch, 2*inch])
        tabla_resumen.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elementos.append(tabla_resumen)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Detalle
        if compras:
            datos_tabla = [['Factura', 'Fecha', 'Proveedor', 'Total']]
            for c in compras:
                fecha = c.get('fecha_compra')
                if isinstance(fecha, str):
                    try:
                        fecha = datetime.strptime(fecha.split()[0], "%Y-%m-%d").date()
                    except:
                        pass
                
                datos_tabla.append([
                    c.get('numero_factura', ''),
                    fecha.strftime("%d/%m/%Y") if hasattr(fecha, 'strftime') else str(fecha),
                    c.get('proveedor', ''),
                    f"Q{c.get('total', 0):,.2f}"
                ])
            
            tabla = Table(datos_tabla, colWidths=[1.5*inch, 1.3*inch, 2.5*inch, 1.2*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#262262')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elementos.append(tabla)
    
    @staticmethod
    def _generar_pdf_productos(elementos, datos, estilo_titulo, estilo_subtitulo):
        """Genera PDF para productos y existencias"""
        stats = datos.get('estadisticas', {})
        productos = datos.get('productos', [])
        
        elementos.append(Paragraph("Productos y Existencias", estilo_titulo))
        elementos.append(Spacer(1, 0.3*inch))
        
        # Estadísticas
        datos_stats = [
            ['Total Productos', str(stats.get('total_productos', 0))],
            ['Sin Stock', str(stats.get('sin_stock', 0))],
            ['Bajo Stock', str(stats.get('bajo_stock', 0))],
            ['Valor Inventario', f"Q{stats.get('valor_inventario', 0):,.2f}"],
        ]
        
        tabla_stats = Table(datos_stats, colWidths=[3*inch, 2*inch])
        tabla_stats.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elementos.append(tabla_stats)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Productos
        if productos:
            datos_tabla = [['Código', 'Producto', 'Stock', 'Mínimo', 'Precio']]
            for p in productos[:50]:  # Limitar a 50 productos
                datos_tabla.append([
                    p.get('codigo', ''),
                    p.get('nombre', '')[:30],
                    str(p.get('stock_actual', 0)),
                    str(p.get('stock_minimo', 0)),
                    f"Q{p.get('precio_venta', 0):,.2f}"
                ])
            
            tabla = Table(datos_tabla, colWidths=[1*inch, 2.5*inch, 0.8*inch, 0.8*inch, 1*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#262262')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8)
            ]))
            elementos.append(tabla)
    
    @staticmethod
    def _generar_pdf_clientes(elementos, datos, estilo_titulo, estilo_subtitulo):
        """Genera PDF para cartera de clientes"""
        stats = datos.get('estadisticas', {})
        clientes = datos.get('clientes', [])
        
        elementos.append(Paragraph("Cartera de Clientes", estilo_titulo))
        elementos.append(Spacer(1, 0.3*inch))
        
        # Stats
        datos_stats = [
            ['Total Clientes', str(stats.get('total_clientes', 0))],
            ['Total Ventas', str(stats.get('total_ventas', 0))],
            ['Ticket Promedio', f"Q{stats.get('ticket_promedio', 0):,.2f}"],
        ]
        
        tabla_stats = Table(datos_stats, colWidths=[3*inch, 2*inch])
        tabla_stats.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elementos.append(tabla_stats)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Clientes
        if clientes:
            datos_tabla = [['Cliente', 'Email', 'Compras', 'Total Gastado']]
            for c in clientes[:40]:
                datos_tabla.append([
                    f"{c.get('nombre', '')} {c.get('apellido', '')}",
                    c.get('email', ''),
                    str(c.get('total_compras', 0)),
                    f"Q{c.get('total_gastado', 0):,.2f}"
                ])
            
            tabla = Table(datos_tabla, colWidths=[2*inch, 2*inch, 1*inch, 1.5*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#262262')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8)
            ]))
            elementos.append(tabla)
    
    @staticmethod
    def _generar_pdf_proveedores(elementos, datos, estilo_titulo, estilo_subtitulo):
        """Genera PDF para cartera de proveedores"""
        stats = datos.get('estadisticas', {})
        proveedores = datos.get('proveedores', [])
        
        elementos.append(Paragraph("Cartera de Proveedores", estilo_titulo))
        elementos.append(Spacer(1, 0.3*inch))
        
        # Stats
        datos_stats = [
            ['Total Proveedores', str(stats.get('total_proveedores', 0))],
            ['Total Compras', str(stats.get('total_compras', 0))],
            ['Compra Promedio', f"Q{stats.get('compra_promedio', 0):,.2f}"],
        ]
        
        tabla_stats = Table(datos_stats, colWidths=[3*inch, 2*inch])
        tabla_stats.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elementos.append(tabla_stats)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Proveedores
        if proveedores:
            datos_tabla = [['Empresa', 'Contacto', 'Compras', 'Total']]
            for p in proveedores[:40]:
                datos_tabla.append([
                    p.get('nombre_empresa', ''),
                    p.get('nombre_contacto', ''),
                    str(p.get('total_compras', 0)),
                    f"Q{p.get('total_comprado', 0):,.2f}"
                ])
            
            tabla = Table(datos_tabla, colWidths=[2*inch, 2*inch, 1*inch, 1.5*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#262262')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8)
            ]))
            elementos.append(tabla)
    
    @staticmethod
    def _generar_pdf_empleados(elementos, datos, estilo_titulo, estilo_subtitulo):
        """Genera PDF para cartera de empleados"""
        stats = datos.get('estadisticas', {})
        empleados = datos.get('empleados', [])
        
        elementos.append(Paragraph("Cartera de Empleados", estilo_titulo))
        elementos.append(Spacer(1, 0.3*inch))
        
        # Stats
        datos_stats = [
            ['Total Empleados', str(stats.get('total_empleados', 0))],
            ['Activos', str(stats.get('empleados_activos', 0))],
            ['Inactivos', str(stats.get('empleados_inactivos', 0))],
        ]
        
        tabla_stats = Table(datos_stats, colWidths=[3*inch, 2*inch])
        tabla_stats.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elementos.append(tabla_stats)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Empleados
        if empleados:
            datos_tabla = [['Empleado', 'Puesto', 'Rol', 'Ventas', 'Estado']]
            for e in empleados[:40]:
                datos_tabla.append([
                    f"{e.get('nombre', '')} {e.get('apellido', '')}",
                    e.get('puesto', ''),
                    e.get('rol', ''),
                    str(e.get('total_ventas_realizadas', 0)),
                    'Activo' if e.get('estado') else 'Inactivo'
                ])
            
            tabla = Table(datos_tabla, colWidths=[1.8*inch, 1.5*inch, 1.2*inch, 0.8*inch, 1*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#262262')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8)
            ]))
            elementos.append(tabla)
    
    @staticmethod
    def exportar_a_excel(datos: Dict[str, Any], tipo_reporte: str, nombre_archivo: str = None) -> str:
        """
        Exporta un reporte a Excel
        
        Args:
            datos: Diccionario con los datos del reporte
            tipo_reporte: Tipo de reporte
            nombre_archivo: Nombre del archivo (opcional)
            
        Returns:
            Ruta del archivo generado
        """
        if not nombre_archivo:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"reporte_{tipo_reporte}_{timestamp}.xlsx"
        
        # Crear directorio de reportes si no existe
        os.makedirs("reportes", exist_ok=True)
        ruta_completa = os.path.join("reportes", nombre_archivo)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Estilos
        titulo_font = Font(name='Arial', size=16, bold=True, color='262262')
        encabezado_fill = PatternFill(start_color='262262', end_color='262262', fill_type='solid')
        encabezado_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Generar contenido según el tipo
        if tipo_reporte == "cierre_diario":
            ExportadorReportes._generar_excel_cierre_diario(ws, datos, titulo_font, encabezado_fill, encabezado_font, border)
        elif tipo_reporte == "cierre_mensual":
            ExportadorReportes._generar_excel_cierre_mensual(ws, datos, titulo_font, encabezado_fill, encabezado_font, border)
        elif tipo_reporte == "compras_periodo":
            ExportadorReportes._generar_excel_compras(ws, datos, titulo_font, encabezado_fill, encabezado_font, border)
        elif tipo_reporte == "productos_existencias":
            ExportadorReportes._generar_excel_productos(ws, datos, titulo_font, encabezado_fill, encabezado_font, border)
        elif tipo_reporte == "cartera_clientes":
            ExportadorReportes._generar_excel_clientes(ws, datos, titulo_font, encabezado_fill, encabezado_font, border)
        elif tipo_reporte == "cartera_proveedores":
            ExportadorReportes._generar_excel_proveedores(ws, datos, titulo_font, encabezado_fill, encabezado_font, border)
        elif tipo_reporte == "cartera_empleados":
            ExportadorReportes._generar_excel_empleados(ws, datos, titulo_font, encabezado_fill, encabezado_font, border)
        
        # Guardar
        wb.save(ruta_completa)
        return ruta_completa
    
    @staticmethod
    def _generar_excel_cierre_diario(ws, datos, titulo_font, encabezado_fill, encabezado_font, border):
        """Genera Excel para cierre diario"""
        ws.title = "Cierre Diario"
        resumen = datos.get('resumen', {})
        ventas = datos.get('ventas', [])
        fecha = datos.get('fecha')
        
        # Título
        ws['A1'] = "CIERRE DE CAJA DIARIO"
        ws['A1'].font = titulo_font
        ws['A2'] = f"Fecha: {fecha.strftime('%d/%m/%Y') if hasattr(fecha, 'strftime') else fecha}"
        
        # Resumen
        ws['A4'] = "RESUMEN"
        ws['A4'].font = Font(bold=True)
        ws['A5'] = "Total Ventas"
        ws['B5'] = resumen.get('total_ventas', 0)
        ws['A6'] = "Total Ingresos"
        ws['B6'] = resumen.get('total_ingresos', 0)
        ws['A7'] = "Efectivo"
        ws['B7'] = resumen.get('efectivo', 0)
        ws['A8'] = "Tarjeta"
        ws['B8'] = resumen.get('tarjeta', 0)
        ws['A9'] = "Transferencia"
        ws['B9'] = resumen.get('transferencia', 0)
        
        # Detalle de ventas
        fila = 11
        ws[f'A{fila}'] = "Factura"
        ws[f'B{fila}'] = "Cliente"
        ws[f'C{fila}'] = "Empleado"
        ws[f'D{fila}'] = "Método Pago"
        ws[f'E{fila}'] = "Total"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{fila}'].fill = encabezado_fill
            ws[f'{col}{fila}'].font = encabezado_font
        
        for v in ventas:
            fila += 1
            ws[f'A{fila}'] = v.get('numero_factura', '')
            ws[f'B{fila}'] = v.get('cliente', '')
            ws[f'C{fila}'] = v.get('empleado', '')
            ws[f'D{fila}'] = v.get('metodo_pago', '').capitalize()
            ws[f'E{fila}'] = v.get('total', 0)
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
    
    @staticmethod
    def _generar_excel_cierre_mensual(ws, datos, titulo_font, encabezado_fill, encabezado_font, border):
        """Genera Excel para cierre mensual"""
        ws.title = "Cierre Mensual"
        resumen = datos.get('resumen', {})
        datos_diarios = datos.get('datos_diarios', [])
        
        meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        
        ws['A1'] = "CIERRE DE CAJA MENSUAL"
        ws['A1'].font = titulo_font
        ws['A2'] = f"{meses[datos['mes']]} {datos['año']}"
        
        # Resumen
        ws['A4'] = "Total Ventas"
        ws['B4'] = resumen.get('total_ventas', 0)
        ws['A5'] = "Total Ingresos"
        ws['B5'] = resumen.get('total_ingresos', 0)
        ws['A6'] = "Promedio por Venta"
        ws['B6'] = resumen.get('promedio_venta', 0)
        
        # Detalle
        fila = 8
        ws[f'A{fila}'] = "Fecha"
        ws[f'B{fila}'] = "Ventas"
        ws[f'C{fila}'] = "Efectivo"
        ws[f'D{fila}'] = "Tarjeta"
        ws[f'E{fila}'] = "Transferencia"
        ws[f'F{fila}'] = "Total"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{fila}'].fill = encabezado_fill
            ws[f'{col}{fila}'].font = encabezado_font
        
        for d in datos_diarios:
            fila += 1
            fecha = d.get('fecha')
            if isinstance(fecha, str):
                try:
                    fecha = datetime.strptime(fecha, "%Y-%m-%d").date()
                except:
                    pass
            
            ws[f'A{fila}'] = fecha.strftime("%d/%m/%Y") if hasattr(fecha, 'strftime') else str(fecha)
            ws[f'B{fila}'] = d.get('total_ventas', 0)
            ws[f'C{fila}'] = d.get('efectivo', 0)
            ws[f'D{fila}'] = d.get('tarjeta', 0)
            ws[f'E{fila}'] = d.get('transferencia', 0)
            ws[f'F{fila}'] = d.get('total_ingresos', 0)
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
    
    @staticmethod
    def _generar_excel_compras(ws, datos, titulo_font, encabezado_fill, encabezado_font, border):
        """Genera Excel para compras"""
        ws.title = "Compras"
        resumen = datos.get('resumen', {})
        compras = datos.get('compras', [])
        
        ws['A1'] = "COMPRAS POR PERIODO"
        ws['A1'].font = titulo_font
        ws['A2'] = f"{datos['fecha_inicio'].strftime('%d/%m/%Y')} - {datos['fecha_fin'].strftime('%d/%m/%Y')}"
        
        # Resumen
        ws['A4'] = "Total Compras"
        ws['B4'] = resumen.get('total_compras', 0)
        ws['A5'] = "Total Gastado"
        ws['B5'] = resumen.get('total_gastado', 0)
        
        # Detalle
        fila = 7
        ws[f'A{fila}'] = "Factura"
        ws[f'B{fila}'] = "Fecha"
        ws[f'C{fila}'] = "Proveedor"
        ws[f'D{fila}'] = "Productos"
        ws[f'E{fila}'] = "Total"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{fila}'].fill = encabezado_fill
            ws[f'{col}{fila}'].font = encabezado_font
        
        for c in compras:
            fila += 1
            fecha = c.get('fecha_compra')
            if isinstance(fecha, str):
                try:
                    fecha = datetime.strptime(fecha.split()[0], "%Y-%m-%d").date()
                except:
                    pass
            
            ws[f'A{fila}'] = c.get('numero_factura', '')
            ws[f'B{fila}'] = fecha.strftime("%d/%m/%Y") if hasattr(fecha, 'strftime') else str(fecha)
            ws[f'C{fila}'] = c.get('proveedor', '')
            ws[f'D{fila}'] = c.get('total_productos', 0)
            ws[f'E{fila}'] = c.get('total', 0)
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
    
    @staticmethod
    def _generar_excel_productos(ws, datos, titulo_font, encabezado_fill, encabezado_font, border):
        """Genera Excel para productos"""
        ws.title = "Productos"
        stats = datos.get('estadisticas', {})
        productos = datos.get('productos', [])
        
        ws['A1'] = "PRODUCTOS Y EXISTENCIAS"
        ws['A1'].font = titulo_font
        
        # Stats
        ws['A3'] = "Total Productos"
        ws['B3'] = stats.get('total_productos', 0)
        ws['A4'] = "Sin Stock"
        ws['B4'] = stats.get('sin_stock', 0)
        ws['A5'] = "Bajo Stock"
        ws['B5'] = stats.get('bajo_stock', 0)
        
        # Detalle
        fila = 7
        ws[f'A{fila}'] = "Código"
        ws[f'B{fila}'] = "Producto"
        ws[f'C{fila}'] = "Categoría"
        ws[f'D{fila}'] = "Stock"
        ws[f'E{fila}'] = "Mínimo"
        ws[f'F{fila}'] = "Precio"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{fila}'].fill = encabezado_fill
            ws[f'{col}{fila}'].font = encabezado_font
        
        for p in productos:
            fila += 1
            ws[f'A{fila}'] = p.get('codigo', '')
            ws[f'B{fila}'] = p.get('nombre', '')
            ws[f'C{fila}'] = p.get('categoria', '')
            ws[f'D{fila}'] = p.get('stock_actual', 0)
            ws[f'E{fila}'] = p.get('stock_minimo', 0)
            ws[f'F{fila}'] = p.get('precio_venta', 0)
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
    
    @staticmethod
    def _generar_excel_clientes(ws, datos, titulo_font, encabezado_fill, encabezado_font, border):
        """Genera Excel para clientes"""
        ws.title = "Clientes"
        stats = datos.get('estadisticas', {})
        clientes = datos.get('clientes', [])
        
        ws['A1'] = "CARTERA DE CLIENTES"
        ws['A1'].font = titulo_font
        
        # Detalle
        fila = 3
        ws[f'A{fila}'] = "Cliente"
        ws[f'B{fila}'] = "Email"
        ws[f'C{fila}'] = "Teléfono"
        ws[f'D{fila}'] = "Compras"
        ws[f'E{fila}'] = "Total Gastado"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{fila}'].fill = encabezado_fill
            ws[f'{col}{fila}'].font = encabezado_font
        
        for c in clientes:
            fila += 1
            ws[f'A{fila}'] = f"{c.get('nombre', '')} {c.get('apellido', '')}"
            ws[f'B{fila}'] = c.get('email', '')
            ws[f'C{fila}'] = c.get('telefono', '')
            ws[f'D{fila}'] = c.get('total_compras', 0)
            ws[f'E{fila}'] = c.get('total_gastado', 0)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
    
    @staticmethod
    def _generar_excel_proveedores(ws, datos, titulo_font, encabezado_fill, encabezado_font, border):
        """Genera Excel para proveedores"""
        ws.title = "Proveedores"
        stats = datos.get('estadisticas', {})
        proveedores = datos.get('proveedores', [])
        
        ws['A1'] = "CARTERA DE PROVEEDORES"
        ws['A1'].font = titulo_font
        
        # Detalle
        fila = 3
        ws[f'A{fila}'] = "Empresa"
        ws[f'B{fila}'] = "Contacto"
        ws[f'C{fila}'] = "Teléfono"
        ws[f'D{fila}'] = "Compras"
        ws[f'E{fila}'] = "Total Comprado"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{fila}'].fill = encabezado_fill
            ws[f'{col}{fila}'].font = encabezado_font
        
        for p in proveedores:
            fila += 1
            ws[f'A{fila}'] = p.get('nombre_empresa', '')
            ws[f'B{fila}'] = p.get('nombre_contacto', '')
            ws[f'C{fila}'] = p.get('telefono', '')
            ws[f'D{fila}'] = p.get('total_compras', 0)
            ws[f'E{fila}'] = p.get('total_comprado', 0)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
    
    @staticmethod
    def _generar_excel_empleados(ws, datos, titulo_font, encabezado_fill, encabezado_font, border):
        """Genera Excel para empleados"""
        ws.title = "Empleados"
        stats = datos.get('estadisticas', {})
        empleados = datos.get('empleados', [])
        
        ws['A1'] = "CARTERA DE EMPLEADOS"
        ws['A1'].font = titulo_font
        
        # Detalle
        fila = 3
        ws[f'A{fila}'] = "Empleado"
        ws[f'B{fila}'] = "Puesto"
        ws[f'C{fila}'] = "Rol"
        ws[f'D{fila}'] = "Teléfono"
        ws[f'E{fila}'] = "Ventas"
        ws[f'F{fila}'] = "Estado"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{fila}'].fill = encabezado_fill
            ws[f'{col}{fila}'].font = encabezado_font
        
        for e in empleados:
            fila += 1
            ws[f'A{fila}'] = f"{e.get('nombre', '')} {e.get('apellido', '')}"
            ws[f'B{fila}'] = e.get('puesto', '')
            ws[f'C{fila}'] = e.get('rol', '')
            ws[f'D{fila}'] = e.get('telefono', '')
            ws[f'E{fila}'] = e.get('total_ventas_realizadas', 0)
            ws[f'F{fila}'] = 'Activo' if e.get('estado') else 'Inactivo'
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
